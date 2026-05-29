import io
import os
from datetime import date as _date
from flask import Flask, jsonify, request, render_template, send_file, abort

import database as db

app = Flask(__name__)
db.init_db()


# ── Helpers ──────────────────────────────────────────────────────────────────

def today_str():
    return _date.today().strftime("%d-%m-%Y")


# ── Page ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# ── Products API ─────────────────────────────────────────────────────────────

@app.route("/api/products", methods=["GET"])
def api_products_list():
    return jsonify(db.get_all_products())


@app.route("/api/products/<dta>", methods=["GET"])
def api_product_get(dta):
    # Try exact match first (case-insensitive)
    product = db.get_product(dta)
    if product:
        return jsonify(product)
        
    # If not found, try digit extraction smart match
    import re
    # Extract only the digit characters from the input search query
    digits = re.sub(r"\D", "", dta)
    if digits:
        all_products = db.get_all_products()
        # Find matches with exact digit sequence matching
        matches = [
            p for p in all_products
            if re.sub(r"\D", "", p["dta"]) == digits
        ]
        # Fallback to contains-digit-sequence matching
        if not matches:
            matches = [
                p for p in all_products
                if digits in re.sub(r"\D", "", p["dta"])
            ]
            
        if matches:
            # Return the first matching product (most relevant)
            return jsonify(matches[0])
            
    return jsonify({"error": "Not found"}), 404


@app.route("/api/products", methods=["POST"])
def api_product_create():
    data = request.get_json()
    if not data or not data.get("dta"):
        return jsonify({"error": "dta is required"}), 400
    db.upsert_product(
        dta=data["dta"],
        brand=data.get("brand", ""),
        model=data.get("model", ""),
        price=data.get("price", 0),
    )
    return jsonify({"ok": True, "dta": data["dta"].upper()})


@app.route("/api/products/<dta>", methods=["PUT"])
def api_product_update(dta):
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400
    db.upsert_product(
        dta=dta,
        brand=data.get("brand", ""),
        model=data.get("model", ""),
        price=data.get("price", 0),
    )
    return jsonify({"ok": True})


@app.route("/api/products/<dta>", methods=["DELETE"])
def api_product_delete(dta):
    db.delete_product(dta)
    return jsonify({"ok": True})

@app.route("/api/products/search", methods=["GET"])
def api_product_search():
    """Return all products whose brand matches AND model prefix matches,
    excluding the current DTA (so you can detect same-product variants).
    Query: ?brand=HP&model=Pavilion+15&exclude_dta=DTA9999
    """
    brand       = (request.args.get("brand") or "").strip().lower()
    model       = (request.args.get("model") or "").strip().lower()
    exclude_dta = (request.args.get("exclude_dta") or "").strip().upper()
    if not brand or not model:
        return jsonify([])
    all_products = db.get_all_products()
    matches = [
        p for p in all_products
        if p["brand"].lower() == brand
        and (
            p["model"].lower().startswith(model[:10])
            or model.startswith(p["model"].lower()[:10])
        )
        and p["dta"].upper() != exclude_dta
    ]
    return jsonify(matches)


def parse_brand_and_model(item_name):
    item_name = item_name.strip()
    if not item_name:
        return "", ""
    
    # Split by whitespace to find the first word (brand)
    parts = item_name.split(None, 1)
    brand = parts[0]
    
    # Keep the brand name in the model field also
    model = item_name
    
    return brand, model


@app.route("/api/products/upload-preview", methods=["POST"])
def api_products_upload_preview():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files["file"]
    filename = file.filename
    if not filename:
        return jsonify({"error": "No file selected"}), 400
    
    ext = os.path.splitext(filename)[1].lower()
    parsed_products = []
    
    try:
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            # Read rows
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(row)
                
        elif ext == ".csv":
            import csv
            # Read CSV
            stream = io.StringIO(file.stream.read().decode("utf-8", errors="ignore"), newline=None)
            csv_reader = csv.reader(stream)
            rows = list(csv_reader)
            
        else:
            return jsonify({"error": "Unsupported file format. Please upload .xlsx or .csv"}), 400
            
        if not rows:
            return jsonify({"error": "The uploaded file is empty"}), 400
            
        # Find headers
        header_row = None
        header_idx = -1
        # Try to find a row containing 'item' and 'name' or similar (using normalized exact matching)
        for idx, row in enumerate(rows[:5]):
            if not row:
                continue
            row_normalized = [str(cell).lower().replace("_", " ").replace("-", " ").strip() if cell is not None else "" for cell in row]
            # Check if any cell matches 'item' or 'dta'
            has_item = any(cell in ["item", "dta", "code", "dta code", "item code"] for cell in row_normalized)
            has_name = any(cell in ["item name", "product name", "model", "description", "product", "name"] for cell in row_normalized)
            if has_item and has_name:
                header_row = row_normalized
                header_idx = idx
                break
                
        # If no clear headers found, assume row 0 is header
        if header_idx == -1:
            header_row = [str(cell).lower().replace("_", " ").replace("-", " ").strip() if cell is not None else "" for cell in rows[0]]
            header_idx = 0
            
        # Map columns
        dta_col_idx = -1
        name_col_idx = -1
        price_col_idx = -1
        brand_col_idx = -1
        
        for idx, val in enumerate(header_row):
            if val in ["dta code", "item code", "dta", "item"]:
                dta_col_idx = idx
            elif val in ["item name", "product name", "model", "description", "product", "name"]:
                name_col_idx = idx
            elif val in ["price", "selling price", "rate", "cost", "value", "unit price"]:
                price_col_idx = idx
            elif val in ["brand", "manufacturer", "make"]:
                brand_col_idx = idx
                
        # Fallbacks if perfect match not found
        if dta_col_idx == -1:
            dta_col_idx = 0
        if name_col_idx == -1:
            name_col_idx = min(1, len(header_row) - 1)
            
        # Parse data rows (start after header_idx)
        for row in rows[header_idx + 1:]:
            if not row or len(row) <= max(dta_col_idx, name_col_idx):
                continue
                
            dta_val = row[dta_col_idx]
            name_val = row[name_col_idx]
            
            if dta_val is None or name_val is None:
                continue
                
            dta = str(dta_val).strip().upper()
            item_name = str(name_val).strip()
            
            if not dta or not item_name:
                continue
                
            # If DTA value is header itself (just in case), skip
            if dta.lower() in ["item", "dta", "code", "dta code", "item name", "product name"]:
                continue
                
            # Parse price if column exists
            price = 0.0
            if price_col_idx != -1 and price_col_idx < len(row):
                p_val = row[price_col_idx]
                if p_val is not None:
                    try:
                        price = float(p_val)
                    except ValueError:
                        pass
                        
            # Parse brand and model
            if brand_col_idx != -1 and brand_col_idx < len(row) and row[brand_col_idx] is not None:
                brand = str(row[brand_col_idx]).strip()
                model = item_name
            else:
                brand, model = parse_brand_and_model(item_name)
                
            parsed_products.append({
                "dta": dta,
                "brand": brand,
                "model": model,
                "price": price
            })
            
        return jsonify({
            "ok": True,
            "filename": filename,
            "total_rows": len(parsed_products),
            "preview": parsed_products[:5],
            "products": parsed_products
        })
        
    except Exception as e:
        return jsonify({"error": f"Error parsing file: {str(e)}"}), 500


@app.route("/api/products/upload-save", methods=["POST"])
def api_products_upload_save():
    data = request.get_json()
    if not data or "products" not in data:
        return jsonify({"error": "No products list provided"}), 400
        
    products = data["products"]
    count = 0
    for p in products:
        dta = (p.get("dta") or "").strip().upper()
        brand = (p.get("brand") or "").strip()
        model = (p.get("model") or "").strip()
        price = float(p.get("price") or 0)
        
        if dta and brand and model:
            db.upsert_product(dta, brand, model, price)
            count += 1
            
    return jsonify({"ok": True, "count": count})




# ── Bills API ─────────────────────────────────────────────────────────────────

@app.route("/api/bills", methods=["GET"])
def api_bills_list():
    date = request.args.get("date", today_str())
    return jsonify(db.get_bills_for_date(date))


@app.route("/api/bills/dates", methods=["GET"])
def api_bill_dates():
    return jsonify(db.get_all_bill_dates())


@app.route("/api/bills", methods=["POST"])
def api_bill_create():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400

    # Auto-save product to DB if DTA is provided
    dta = (data.get("dta") or "").strip().upper()
    if dta and data.get("brand") and data.get("model"):
        db.upsert_product(dta, data["brand"], data["model"], data.get("price", 0))

    # For exchange: also save new product DTA
    exch_dta = (data.get("exch_new_dta") or "").strip().upper()
    if exch_dta and data.get("exch_new_brand") and data.get("exch_new_model"):
        db.upsert_product(
            exch_dta,
            data["exch_new_brand"],
            data["exch_new_model"],
            data.get("exch_new_price", 0),
        )

    if not data.get("date"):
        data["date"] = today_str()

    new_id = db.create_bill(data)
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/bills/<int:bill_id>", methods=["DELETE"])
def api_bill_delete(bill_id):
    db.delete_bill(bill_id)
    return jsonify({"ok": True})


# ── Stats API ─────────────────────────────────────────────────────────────────

@app.route("/api/stats", methods=["GET"])
def api_stats():
    return jsonify(db.get_stats())


# ── Export API ────────────────────────────────────────────────────────────────

@app.route("/api/export", methods=["GET"])
def api_export():
    date = request.args.get("date", today_str())
    bills = db.get_bills_for_date(date)

    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sales {date}"

    # ── Colours ───────────────────────────────────────────────────────────────
    yellow_fill   = PatternFill("solid", fgColor="FFD700")
    blue_fill     = PatternFill("solid", fgColor="1E3A5F")
    alt_fill      = PatternFill("solid", fgColor="EBF2FF")
    exch_fill     = PatternFill("solid", fgColor="DDEEFF")   # light blue for exchange rows
    collect_fill  = PatternFill("solid", fgColor="D4EDDA")   # green for positive balance
    refund_fill   = PatternFill("solid", fgColor="F8D7DA")   # red for negative balance

    white_font    = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    header_font   = Font(bold=True, size=13, name="Calibri")
    bold_font     = Font(bold=True, name="Calibri")
    green_font    = Font(bold=True, color="155724", name="Calibri")
    red_font      = Font(bold=True, color="721C24", name="Calibri")
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    COLS = [
        "#", "Customer", "Brand", "Model", "DTA", "Price (AED)",
        "MOP", "Note", "Type", "Platform", "Delivery", "Mixed Detail",
        # Exchange columns
        "Exch — New Product", "Exch — New DTA", "Exch — New Price (AED)",
        "Exch — Old Product", "Exch — Old DTA", "Exch — Old Price (AED)",
        "Exch — Balance (AED)",
    ]
    NUM_COLS = len(COLS)

    # ── Row 1: Date header ────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    cell = ws.cell(row=1, column=1, value=f"Sales Sheet — {date}")
    cell.fill = yellow_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, bill in enumerate(bills, 1):
        row_num = ri + 2
        is_exchange = bill["transaction_type"] == "Exchange"
        row_fill = exch_fill if is_exchange else (alt_fill if ri % 2 == 0 else None)

        mixed_detail = ""
        if bill["payment_mode"] == "Mixed":
            parts = []
            if bill["mixed_cash"]:   parts.append(f"Cash: AED {bill['mixed_cash']:.0f}")
            if bill["mixed_card"]:   parts.append(f"Card: AED {bill['mixed_card']:.0f}")
            if bill["mixed_tabby"]:  parts.append(f"Tabby: AED {bill['mixed_tabby']:.0f}")
            if bill["mixed_tamara"]: parts.append(f"Tamara: AED {bill['mixed_tamara']:.0f}")
            mixed_detail = " | ".join(parts)

        # For exchange, show new product in main brand/model/dta/price columns
        brand = bill["exch_new_brand"] or bill["brand"] if is_exchange else bill["brand"]
        model = bill["exch_new_model"] or bill["model"] if is_exchange else bill["model"]
        dta   = bill["exch_new_dta"]   or bill["dta"]   if is_exchange else bill["dta"]
        price = bill["exch_new_price"] or bill["price"] if is_exchange else bill["price"]

        # Exchange detail columns
        exch_new_product = f"{bill['exch_new_brand']} {bill['exch_new_model']}".strip() if is_exchange else ""
        exch_new_dta     = bill["exch_new_dta"] if is_exchange else ""
        exch_new_price   = bill["exch_new_price"] if is_exchange else ""
        exch_old_product = f"{bill['exch_old_brand']} {bill['exch_old_model']}".strip() if is_exchange else ""
        exch_old_dta     = bill["exch_old_dta"] if is_exchange else ""
        exch_old_price   = bill["exch_old_price"] if is_exchange else ""
        exch_balance     = bill["exch_balance"] if is_exchange else ""

        row_data = [
            ri,
            bill["customer_name"],
            brand,
            model,
            dta,
            price,
            bill["payment_mode"],
            bill["note"],
            bill["transaction_type"],
            bill["platform"],
            "Yes" if bill["delivery"] else "No",
            mixed_detail,
            exch_new_product,
            exch_new_dta,
            exch_new_price,
            exch_old_product,
            exch_old_dta,
            exch_old_price,
            exch_balance,
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill

        # Color-code the balance cell (column 19)
        if is_exchange and exch_balance != "":
            bal_cell = ws.cell(row=row_num, column=19)
            if isinstance(exch_balance, (int, float)):
                if exch_balance > 0:
                    bal_cell.fill   = collect_fill
                    bal_cell.font   = green_font
                    bal_cell.value  = f"+{exch_balance:.2f}  ↑ Collect"
                elif exch_balance < 0:
                    bal_cell.fill   = refund_fill
                    bal_cell.font   = red_font
                    bal_cell.value  = f"{exch_balance:.2f}  ↓ Refund"
                else:
                    bal_cell.font   = bold_font
                    bal_cell.value  = "0.00  ✓ Even"

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [4, 18, 14, 18, 12, 14, 10, 26, 10, 18, 9, 28,
              22, 12, 16, 22, 12, 16, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # Freeze top 2 rows
    ws.freeze_panes = "A3"

    # ── Save to buffer ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Sales_{date}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app.run(debug=True, port=5000)

